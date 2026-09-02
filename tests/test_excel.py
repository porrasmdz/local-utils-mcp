import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from services import excel


def _write_minimal_xlsx(path: Path):
    files = {
        "[Content_Types].xml": """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
</Types>""",
        "xl/workbook.xml": """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Summary" sheetId="1" r:id="rId1"/>
    <sheet name="HiddenData" sheetId="2" state="hidden" r:id="rId2"/>
  </sheets>
</workbook>""",
        "xl/_rels/workbook.xml.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Target="worksheets/sheet2.xml"/>
</Relationships>""",
        "xl/sharedStrings.xml": """<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <si><t>Name</t></si>
  <si><t>Total</t></si>
  <si><t>Alice</t></si>
  <si><t>Bob</t></si>
</sst>""",
        "xl/worksheets/sheet1.xml": """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheetData>
    <row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>1</v></c></row>
    <row r="2"><c r="A2" t="s"><v>2</v></c><c r="B2"><v>10</v></c></row>
    <row r="3"><c r="A3" t="s"><v>3</v></c><c r="B3"><f>SUM(B2:B2)</f><v>10</v></c></row>
  </sheetData>
  <mergeCells count="1"><mergeCell ref="A1:B1"/></mergeCells>
  <hyperlinks><hyperlink ref="A2" r:id="rId1"/></hyperlinks>
  <tableParts count="1"><tablePart r:id="rId2"/></tableParts>
  <drawing r:id="rId3"/>
</worksheet>""",
        "xl/worksheets/sheet2.xml": """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData/>
</worksheet>""",
    }

    with zipfile.ZipFile(path, "w") as zf:
        for name, content in files.items():
            zf.writestr(name, content)


class ExcelSummaryTest(unittest.TestCase):
    def test_summarize_excel_workbook_returns_sheets_ranges_objects_and_preview(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "workbook.xlsx"
            _write_minimal_xlsx(path)

            result = excel.summarize_excel_workbook(str(path))

        self.assertIsInstance(result, excel.ExcelWorkbookSummary)
        self.assertEqual(result.sheet_count, 2)
        self.assertEqual(result.sheets[0].sheet_name, "Summary")
        self.assertEqual(result.sheets[0].sheet_state, "visible")
        self.assertEqual(result.sheets[0].used_range, "A1:B3")
        self.assertEqual(result.sheets[0].first_content_row, 1)
        self.assertEqual(result.sheets[0].last_content_row, 3)
        self.assertEqual(result.sheets[0].first_10_rows, [
            ["Name", "Total"],
            ["Alice", 10],
            ["Bob", 10],
        ])
        self.assertEqual(result.sheets[0].objects, {
            "tables": 1,
            "drawings": 1,
            "hyperlinks": 1,
            "merged_ranges": 1,
            "comments": 0,
            "formulas": 1,
        })
        self.assertEqual(result.sheets[1].sheet_name, "HiddenData")
        self.assertEqual(result.sheets[1].sheet_state, "hidden")
        self.assertIsNone(result.sheets[1].used_range)
        self.assertEqual(result.sheets[1].first_10_rows, [])

    def test_summarize_excel_workbook_rejects_unsupported_extension(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "workbook.csv"
            path.write_text("a,b\n1,2\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, ".xlsx"):
                excel.summarize_excel_workbook(str(path))

    def test_summarize_excel_workbook_rejects_missing_file(self):
        with self.assertRaises(FileNotFoundError):
            excel.summarize_excel_workbook("missing.xlsx")


class ExcelRowsTest(unittest.TestCase):
    def test_read_excel_sheet_rows_returns_requested_range(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "workbook.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(["Name", "Total", "Ignored"])
            sheet.append(["Alice", 10, "x"])
            sheet.append(["Bob", 20, "y"])
            workbook.save(path)

            result = excel.read_excel_sheet_rows(
                str(path),
                "Data",
                start_row=2,
                end_row=3,
                min_column=1,
                max_column=2,
            )

        self.assertIsInstance(result, excel.ExcelRowsResult)
        self.assertEqual(result.sheet_name, "Data")
        self.assertEqual(result.start_row, 2)
        self.assertEqual(result.end_row, 3)
        self.assertEqual(result.min_column, 1)
        self.assertEqual(result.max_column, 2)
        self.assertEqual([row.row_number for row in result.rows], [2, 3])
        self.assertEqual([row.values for row in result.rows], [["Alice", 10], ["Bob", 20]])

    def test_read_excel_sheet_rows_rejects_missing_sheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "workbook.xlsx"
            workbook = Workbook()
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "No existe la hoja"):
                excel.read_excel_sheet_rows(str(path), "Missing", 1, 2)


class ExcelUpdateTest(unittest.TestCase):
    def test_update_excel_sheet_cells_writes_cells_and_merges_ranges_centered(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "workbook.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            workbook.save(path)

            result = excel.update_excel_sheet_cells(
                str(path),
                "Report",
                {
                    "A1": "Title",
                    "B2:C2": "Merged Value",
                },
            )

            workbook = load_workbook(path)
            sheet = workbook["Report"]
            try:
                self.assertIsInstance(result, excel.ExcelUpdateResult)
                self.assertEqual(
                    [(item.target, item.value, item.merged) for item in result.updated_cells],
                    [("A1", "Title", False), ("B2:C2", "Merged Value", True)],
                )
                self.assertEqual(sheet["A1"].value, "Title")
                self.assertEqual(sheet["B2"].value, "Merged Value")
                self.assertIn("B2:C2", [str(item) for item in sheet.merged_cells.ranges])
                self.assertEqual(sheet["B2"].alignment.horizontal, "center")
                self.assertEqual(sheet["B2"].alignment.vertical, "center")
            finally:
                workbook.close()

    def test_update_excel_sheet_cells_rejects_invalid_target(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "workbook.xlsx"
            workbook = Workbook()
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "Referencia"):
                excel.update_excel_sheet_cells(str(path), "Sheet", {"not-a-cell": "x"})


if __name__ == "__main__":
    unittest.main()
